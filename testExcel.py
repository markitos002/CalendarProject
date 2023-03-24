import openpyxl
import calendar

#create a workbook

wb = openpyxl.Workbook()

# select the active sheet

sheet = wb.active

style = openpyxl.styles.Font(bold=False, size=12)


# write data to the sheet

datos = [['name', 'age', 'gender'], ['John', 25, 'male'], ['Mary', 30, 'female'], ['Peter', 20 , 'other'], ['Susan', 33, 'non-binary']]
print(datos)
for i in range(len(datos)):
    for j in range(len(datos[i])):
        sheet.cell(row=i+1, column=j+1).value = datos[i][j]
        sheet.cell(row=i+1, column=j+1).font = style
       
            
"""
sheet['A1'] = "name"
sheet['B1'] = "age"
sheet['A2'] = "John"
sheet['B2'] = 25
sheet['A3'] = "Mary"
sheet['B3'] = 30
"""
# save the file

wb.save("test.xlsx")


