
import calendar
from openpyxl import Workbook
"""
# Set the year and range of months to generate the calendar for
year = 2023
start_month = 3
end_month = 9

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Loop through the months and write the calendar data to the worksheet
for month in range(start_month, end_month+1):
    # Get the month name and header for the calendar
    month_name = calendar.month_name[month]
    month_calendar = calendar.monthcalendar(year, month)
    month_header = f"{month_name} {year}"

    # Write the month header and calendar to the worksheet
    col_start = (month - start_month) * 22 + 1
    ws.cell(row=1, column=col_start, value=month_header)
    for i, day in enumerate(["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"], start=col_start):
        ws.cell(row=2, column=i, value=day)
    for r, week in enumerate(month_calendar, 3):
        for c, day in enumerate(week):
            if day != 0:
                ws.cell(row=r, column=col_start+c, value=day)

# Save the workbook to a file
wb.save('calendar.xlsx')
"""
import calendar
from openpyxl import Workbook

# Set the year and range of months to generate the calendar for
year = 2023
start_month = 3
end_month = 9

# Create a new workbook
wb = Workbook()

# Loop through the months and add a worksheet for each one
for month in range(start_month, end_month+1):
    # Get the month name and header for the calendar
    month_name = calendar.month_name[month]
    month_calendar = calendar.monthcalendar(year, month)
    month_header = f"{month_name} {year}"

    # Create a new worksheet for the month
    ws = wb.create_sheet(month_name)

    # Write the month header and calendar to the worksheet
    ws.cell(row=1, column=1, value=month_header)
    ws.cell(row=2, column=1, value="Su Mo Tu We Th Fr Sa")
    for r, week in enumerate(month_calendar, 3):
        for c, day in enumerate(week):
            ws.cell(row=r, column=c+1, value=day)

# Save the workbook to a file
wb.save('calendar.xlsx')
"""

import calendar

# Set the year and range of months to generate the calendar for
year = 2023
start_month = 3
end_month = 9

# Loop through the months and generate the calendar for each one
for month in range(start_month, end_month+1):
    # Get the month name and header for the calendar
    month_name = calendar.month_name[month]
    month_calendar = calendar.monthcalendar(year, month)
    month_header = f"{' '*(20-len(month_name)//2)}{month_name} {year}{' '*(20-len(month_name)//2)}\n{'-'*44}\n{'Su Mo Tu We Th Fr Sa'}"

    # Print the month header and calendar
    print(month_header)
    for week in month_calendar:
        week_str = ' '.join([str(day).rjust(2) if day != 0 else '  ' for day in week])
        print(week_str)
    print('')

"""