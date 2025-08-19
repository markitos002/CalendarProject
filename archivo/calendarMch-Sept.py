import calendar
import openpyxl

# create a new workbook
workbook = openpyxl.Workbook()

# select the active worksheet
worksheet = workbook.active

# set up the header row with the weekday names
header = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
for i, day in enumerate(header):
    cell = worksheet.cell(row=1, column=i+1)
    cell.value = day

# get the calendar data for the desired date range
start_date = (2023, 3, 20)
end_date = (2023, 9, 1)
cal = calendar.monthcalendar(start_date[0], start_date[1])[:6] + calendar.monthcalendar(end_date[0], end_date[1])[:6]

# write the calendar data to the worksheet

row_num = 2
for month in cal:
    for week in month:
        for i, day in enumerate(week):
            cell = worksheet.cell(row=row_num, column=i+1)
            if day.month == 3:
                cell.value = f"{day.day}-Mar"
            elif day.month == 9:
                cell.value = f"{day.day}-Sep"
            else:
                cell.value = day.day
            if day == calendar.datetime.date.today():
                cell.font = openpyxl.styles.Font(bold=True, color="FF0000") # highlight today's date in red
        row_num += 1

# save the workbook

workbook.save("calendarMch-Sept.xlsx")
