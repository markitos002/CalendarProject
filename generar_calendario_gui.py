"""
Generador de calendario de clases (GUI)

Resumen de arquitectura
- Capa de datos:
    - WeekDates: estructura con las fechas de cada semana (Lun/Mar/Mié) y el número de semana.
    - compute_weeks(): calcula 18 semanas (o N) a partir de un lunes de inicio.
    - get_colombia_holidays(): obtiene festivos en Colombia para el rango [inicio, fin]; usa la
        librería "holidays" si está disponible, o un fallback mínimo para 2025.

- Capa de exportación:
    - build_excel(): genera un archivo .xlsx con una tabla por semana (encabezado + 4 columnas).
    - build_pdf(): genera un PDF con tablas por semana (opcional; requiere reportlab).

- Capa de presentación (GUI):
    - CalendarGUI (Tkinter): ofrece controles para título/subtítulo, fecha de inicio (con tkcalendar
        si está instalado) y número de semanas; muestra una grilla editable por semana y exporta a Excel/PDF.

Notas de mantenimiento
- Si en el futuro se agregan más días (p. ej. Jueves/Viernes), modifica:
    1) WeekDates (añadir campos o cambiar el diseño).
    2) _build_weeks_ui() para crear columnas y textos.
    3) build_excel()/build_pdf() para reflejar las nuevas columnas.
    4) _collect_entries() para devolver los nuevos textos.
"""

import sys
import os
import json
from datetime import date, timedelta
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple, Set

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except Exception:
    tk = None  # type: ignore

# Optional date picker (nice UX if available)
try:
    from tkcalendar import DateEntry  # type: ignore
    TKCAL_OK = True
except Exception:
    TKCAL_OK = False

# Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Optional PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

# Optional Holidays (for Colombia)
def _fallback_colombia_holidays_2025() -> Dict[date, str]:
    """Minimal fallback for 2025 Colombian holidays that affect Aug–Dec period.

    This covers Monday-observed holidays within the semester window.
    """
    return {
        date(2025, 8, 18): "Asunción de la Virgen (festivo)",
        date(2025, 10, 13): "Día de la Raza (festivo)",
        date(2025, 11, 3): "Día de Todos los Santos (festivo)",
        date(2025, 11, 17): "Independencia de Cartagena (festivo)",
        date(2025, 12, 8): "Inmaculada Concepción (festivo)",
    }


def get_colombia_holidays(start: date, end: date) -> Dict[date, str]:
    """Return a dict of holiday_date -> holiday_name for Colombia within range.

    Tries the 'holidays' package; falls back to a minimal 2025 set.
    """
    try:
        import holidays  # type: ignore

        co = holidays.country_holidays("CO", years={start.year, end.year})
        out: Dict[date, str] = {}
        d = start
        while d <= end:
            if d in co:
                out[d] = str(co.get(d))
            d += timedelta(days=1)
        return out
    except Exception:
        fallback = _fallback_colombia_holidays_2025()
        return {k: v for k, v in fallback.items() if start <= k <= end}


SPANISH_MONTHS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


@dataclass
class WeekDates:
    semana: int
    lunes: date
    martes: date
    miercoles: date


def compute_weeks(start_monday: date, weeks: int = 18) -> List[WeekDates]:
    """Calcula un arreglo de WeekDates a partir de un lunes de inicio.

    Parámetros
    - start_monday: fecha que debe ser lunes (weekday()==0).
    - weeks: cantidad de semanas a generar.

    Retorna
    - Lista de WeekDates con (semana, lunes, martes, miércoles) por cada semana.
    """
    if start_monday.weekday() != 0:
        raise ValueError("La fecha de inicio debe ser un lunes")
    out: List[WeekDates] = []
    for i in range(weeks):
        mon = start_monday + timedelta(weeks=i)
        out.append(WeekDates(semana=i + 1, lunes=mon, martes=mon + timedelta(days=1), miercoles=mon + timedelta(days=2)))
    return out


def build_excel(
    out_path: str,
    title: str,
    subtitle: str,
    week_dates: List[WeekDates],
    entries: Dict[int, Tuple[str, str, str, str]],
    holidays_map: Dict[date, str],
    exam_dates: Set[date],
) -> None:
    """Crea un archivo Excel con el calendario.

        Estructura de salida
        - Título y subtítulo en las filas 1 y 2 (celdas combinadas).
        - Por cada semana: una fila de encabezado "SEMANA X MES" y debajo la fila con 4 columnas
            (Lunes, Martes, Miércoles1, Miércoles2). Si un día es festivo, se rellena la celda y se
            escribe "No hay clase".
        - Deja una fila en blanco entre semanas para mejorar la legibilidad.

        Notas
        - Usa estilos simples (bordes finos, rellenos y alineaciones) para facilitar cambios futuros.
        """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"

    # Styles
    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=False, size=12)
    week_header_font = Font(bold=True, size=12)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True, vertical="top")
    # Colors: holidays light green, exams light orange
    holiday_fill = PatternFill("solid", fgColor="C6EFCE")
    exam_fill = PatternFill("solid", fgColor="F8CBAD")

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1, value=title).font = header_font
    ws.cell(row=1, column=1).alignment = align_center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1, value=subtitle).font = sub_font
    ws.cell(row=2, column=1).alignment = align_center

    row = 4
    # Column widths
    widths = [22, 22, 22, 22, 1, 1]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for wd in week_dates:
        month_name = SPANISH_MONTHS[wd.lunes.month]
        # Week header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"SEMANA {wd.semana} {month_name}").font = week_header_font
        ws.cell(row=row, column=1).alignment = align_center
        row += 1

        # Day headers
        headers = [
            f"Lunes {wd.lunes.strftime('%d/%m')}",
            f"Martes {wd.martes.strftime('%d/%m')}",
            f"Miércoles 1 {wd.miercoles.strftime('%d/%m')}",
            f"Miércoles 2 {wd.miercoles.strftime('%d/%m')}",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True)
            c.alignment = align_center
            c.border = border
        row += 1

        # Content row
        mon_txt, tue_txt, wed1_txt, wed2_txt = entries.get(wd.semana, ("", "", "", ""))

        for col, (d, txt) in enumerate(
            [
                (wd.lunes, mon_txt),
                (wd.martes, tue_txt),
                (wd.miercoles, wed1_txt),
                (wd.miercoles, wed2_txt),
            ],
            start=1,
        ):
            cell = ws.cell(row=row, column=col)
            is_holiday = d in holidays_map
            if is_holiday:
                cell.value = f"Festivo: {holidays_map[d]}\nNo hay clase"
                cell.fill = holiday_fill
            elif d in exam_dates:
                # Mark exams with yellow fill and optional text
                cell.value = f"Examen" + (f"\n{txt}" if txt else "")
                cell.fill = exam_fill
            else:
                cell.value = txt
            cell.alignment = align_wrap
            cell.border = border
        row += 2  # leave a blank row between weeks

    wb.save(out_path)


def build_pdf(
    out_path: str,
    title: str,
    subtitle: str,
    week_dates: List[WeekDates],
    entries: Dict[int, Tuple[str, str, str, str]],
    holidays_map: Dict[date, str],
    exam_dates: Set[date],
) -> None:
    """Crea un PDF con el calendario por tablas (opcional).

    - Cada semana se imprime como una tabla de 2 filas: encabezados (días) y contenidos.
    - Sombrea las celdas de días festivos para diferenciarlas.
    - Requiere la librería reportlab. Si no está, se lanza un RuntimeError controlado.
    """
    if not REPORTLAB_OK:
        raise RuntimeError("ReportLab no está instalado. Instálalo para exportar a PDF.")

    doc = SimpleDocTemplate(out_path, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    parts: List = []
    parts.append(Paragraph(title, styles["Title"]))
    parts.append(Paragraph(subtitle, styles["Normal"]))
    parts.append(Spacer(1, 10))

    for wd in week_dates:
        month_name = SPANISH_MONTHS[wd.lunes.month]
        parts.append(Paragraph(f"SEMANA {wd.semana} {month_name}", styles["Heading2"]))

        data = [
            [
                f"Lunes {wd.lunes.strftime('%d/%m')}",
                f"Martes {wd.martes.strftime('%d/%m')}",
                f"Miércoles 1 {wd.miercoles.strftime('%d/%m')}",
                f"Miércoles 2 {wd.miercoles.strftime('%d/%m')}",
            ]
        ]

        mon_txt, tue_txt, wed1_txt, wed2_txt = entries.get(wd.semana, ("", "", "", ""))

        def cell_text(d: date, txt: str) -> str:
            if d in holidays_map:
                return f"Festivo: {holidays_map[d]}\nNo hay clase"
            if d in exam_dates:
                return "Examen" + (f"\n{txt}" if txt else "")
            return txt

        data.append([
            cell_text(wd.lunes, mon_txt),
            cell_text(wd.martes, tue_txt),
            cell_text(wd.miercoles, wed1_txt),
            cell_text(wd.miercoles, wed2_txt),
        ])

        t = Table(data, colWidths=[200, 200, 200, 200])
        t.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        # Shade holiday/exam cells
        holidays_cols = []
        if wd.lunes in holidays_map:
            holidays_cols.append(0)
        if wd.martes in holidays_map:
            holidays_cols.append(1)
        if wd.miercoles in holidays_map:
            holidays_cols.extend([2, 3])
        for c in holidays_cols:
            # Light green for holidays
            t.setStyle(TableStyle([("BACKGROUND", (c, 1), (c, 1), colors.HexColor('#C6EFCE'))]))

        exam_cols = []
        if wd.lunes in exam_dates:
            exam_cols.append(0)
        if wd.martes in exam_dates:
            exam_cols.append(1)
        if wd.miercoles in exam_dates:
            exam_cols.extend([2, 3])
        for c in exam_cols:
            # Light orange for exams
            t.setStyle(TableStyle([("BACKGROUND", (c, 1), (c, 1), colors.HexColor('#F8CBAD'))]))

        parts.append(t)
        parts.append(Spacer(1, 6))

    doc.build(parts)


class CalendarGUI:
    """Ventana principal de la aplicación.

    Responsabilidades
    - Proveer controles de entrada (título, subtítulo, fecha de inicio, semanas).
    - Renderizar la grilla editable por semana con 4 columnas (Lu, Ma, Mié1, Mié2).
    - Deshabilitar celdas de días festivos con el mensaje "No hay clase".
    - Exportar a Excel/PDF con los datos ingresados.
    """
    def __init__(self, root: tk.Tk, start: date = date(2025, 8, 18), weeks: int = 18) -> None:  # type: ignore[name-defined]
        self.root = root
        self.root.title("Generador de Calendario de Clases")
        self.start = start
        self.weeks = weeks
        self.week_dates = compute_weeks(start, weeks)
        self.end = self.week_dates[-1].miercoles
        self.holidays = get_colombia_holidays(self.start, self.end)

        # Top fields
        top = ttk.Frame(root)
        top.pack(fill=tk.X, padx=10, pady=(10, 4))

        ttk.Label(top, text="Título (curso)").grid(row=0, column=0, sticky=tk.W)
        self.var_title = tk.StringVar(value="Fundamentos de Ciencias Básicas - 2025 - B")
        ttk.Entry(top, width=60, textvariable=self.var_title).grid(row=0, column=1, sticky=tk.W)

        ttk.Label(top, text="Subtítulo (opcional)").grid(row=1, column=0, sticky=tk.W)
        self.var_sub = tk.StringVar(value=f"Desde {self.start.strftime('%d/%m/%Y')} por {self.weeks} semanas")
        ttk.Entry(top, width=60, textvariable=self.var_sub).grid(row=1, column=1, sticky=tk.W)

        # Controls: start date + weeks + update
        controls = ttk.Frame(root)
        controls.pack(fill=tk.X, padx=10, pady=(0, 6))
        ttk.Label(controls, text="Fecha de inicio (debe ser Lunes):").grid(row=0, column=0, sticky=tk.W)

        if TKCAL_OK:
            self.date_widget = DateEntry(controls, date_pattern='dd/mm/yyyy')
            self.date_widget.set_date(self.start)
            self.date_widget.grid(row=0, column=1, padx=(6, 12))
        else:
            # Fallback: comboboxes for day/month/year
            self._var_day = tk.StringVar(value=str(self.start.day))
            self._var_month = tk.StringVar(value=str(self.start.month))
            self._var_year = tk.StringVar(value=str(self.start.year))
            ttk.Combobox(controls, width=3, values=[f"{i:02d}" for i in range(1, 32)], textvariable=self._var_day, state="readonly").grid(row=0, column=1, padx=3)
            ttk.Combobox(controls, width=3, values=[f"{i:02d}" for i in range(1, 13)], textvariable=self._var_month, state="readonly").grid(row=0, column=2, padx=3)
            ttk.Combobox(controls, width=5, values=[str(y) for y in range(self.start.year-2, self.start.year+3)], textvariable=self._var_year, state="readonly").grid(row=0, column=3, padx=(3, 12))

        ttk.Label(controls, text="Semanas:").grid(row=0, column=4, sticky=tk.W)
        self.var_weeks = tk.StringVar(value=str(self.weeks))
        ttk.Spinbox(controls, from_=1, to=30, width=4, textvariable=self.var_weeks).grid(row=0, column=5, padx=(6, 12))
        ttk.Button(controls, text="Actualizar calendario", command=self.rebuild_calendar).grid(row=0, column=6)

        # Exams input (8 dates)
        exams_frame = ttk.LabelFrame(root, text="Fechas de exámenes (8)")
        exams_frame.pack(fill=tk.X, padx=10, pady=(0, 6))
        self.exam_inputs: List[Any] = []
        if TKCAL_OK:
            for i in range(8):
                ttk.Label(exams_frame, text=f"Examen {i+1}:").grid(row=i//4, column=(i%4)*2, sticky=tk.W, padx=(6 if i%4 else 10, 4), pady=2)
                de = DateEntry(exams_frame, date_pattern='dd/mm/yyyy')
                de.grid(row=i//4, column=(i%4)*2 + 1, padx=(0, 12), pady=2)
                self.exam_inputs.append(de)
        else:
            # Fallback: 8 Entry fields dd/mm/yyyy
            for i in range(8):
                ttk.Label(exams_frame, text=f"Examen {i+1} (dd/mm/yyyy):").grid(row=i//4, column=(i%4)*2, sticky=tk.W, padx=(6 if i%4 else 10, 4), pady=2)
                ent = ttk.Entry(exams_frame, width=12)
                ent.grid(row=i//4, column=(i%4)*2 + 1, padx=(0, 12), pady=2)
                self.exam_inputs.append(ent)

        # Contenedor con Canvas + Scrollbar para poder desplazar 18 filas cómodamente
        container = ttk.Frame(root)
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        canvas = tk.Canvas(container, height=480)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.scroll_frame = ttk.Frame(canvas)
        self.scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Build weeks UI initially
        self._build_weeks_ui()

        # Action buttons
        actions = ttk.Frame(root)
        actions.pack(fill=tk.X, padx=10, pady=(0, 12))
        ttk.Button(actions, text="Exportar a Excel (.xlsx)", command=self.export_excel).pack(side=tk.LEFT)
        ttk.Button(actions, text="Exportar a PDF (.pdf)", command=self.export_pdf).pack(side=tk.LEFT, padx=10)

        # Holidays notice
        self.info_label = ttk.Label(root, text=self._holidays_text())
        self.info_label.pack(padx=10, pady=(0, 10))

        # Hook close event to save backup
        try:
            self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        except Exception:
            pass

        # Load previous backup and apply
        self._load_backup()

    def _get_exam_dates(self) -> Set[date]:
        """Lee y valida las 8 fechas de exámenes del UI; ignora vacíos."""
        out: Set[date] = set()
        for widget in self.exam_inputs:
            try:
                if TKCAL_OK and hasattr(widget, 'get_date'):
                    d = widget.get_date()
                    out.add(date(d.year, d.month, d.day))
                else:
                    txt = widget.get().strip()
                    if not txt:
                        continue
                    # Parse dd/mm/yyyy
                    dd, mm, yyyy = txt.split('/')
                    out.add(date(int(yyyy), int(mm), int(dd)))
            except Exception:
                # Ignore malformed entries
                continue
        return out

    def _holidays_text(self) -> str:
        """Construye el texto de resumen de festivos actual."""
        if not self.holidays:
            return "Festivos detectados: ninguno en el rango"
        return f"Festivos detectados: {', '.join(d.strftime('%d/%m') for d in sorted(self.holidays.keys()))}"

    def _get_selected_start_date(self) -> date:
        """Lee la fecha desde tkcalendar o desde los Combobox de fallback."""
        if TKCAL_OK:
            d = self.date_widget.get_date()
            # tkcalendar returns datetime.date
            return date(d.year, d.month, d.day)
        # Fallback from comboboxes
        y = int(self._var_year.get())
        m = int(self._var_month.get())
        d = int(self._var_day.get())
        return date(y, m, d)

    def rebuild_calendar(self) -> None:
        """Recalcula semanas y festivos cuando cambia inicio/semanas.

        Pasos:
        1) Leer fecha y número de semanas.
        2) Validar que el inicio sea lunes.
        3) Recalcular self.week_dates y self.holidays.
        4) Destruir la grilla anterior y volver a construirla.
        """
        try:
            start = self._get_selected_start_date()
        except Exception as e:
            messagebox.showerror("Fecha inválida", f"No se pudo leer la fecha: {e}")
            return
        try:
            weeks = int(self.var_weeks.get())
        except Exception:
            weeks = 18

        if start.weekday() != 0:
            messagebox.showerror("Inicio inválido", "La fecha de inicio debe ser un Lunes.")
            return

        self.start = start
        self.weeks = weeks
        self.week_dates = compute_weeks(self.start, self.weeks)
        self.end = self.week_dates[-1].miercoles
        self.holidays = get_colombia_holidays(self.start, self.end)

        # Rebuild scroll area
        for child in self.scroll_frame.winfo_children():
            child.destroy()
        self._build_weeks_ui()
        self.info_label.config(text=self._holidays_text())
        # After rebuilding, re-apply saved entries if any
        saved = self._read_backup_file()
        if saved and isinstance(saved.get("entries"), dict):
            self._apply_saved_entries(saved["entries"])

    def _build_weeks_ui(self) -> None:
        """Construye la grilla de semanas sobre un único grid.

        Diseño:
        - Fila 0: Encabezados (Semana, Lunes, Martes, Mié1, Mié2) colocados en ``self.scroll_frame``.
        - Filas 1..N: Por cada semana, una etiqueta con el rango y 4 widgets de texto.
        - Importante: se utiliza el mismo contenedor y las mismas columnas para mantener
          alineados encabezados y contenido (evita desfases).
        """
        for c in range(5):
            self.scroll_frame.grid_columnconfigure(c, weight=0)

        # Headers directly in scroll_frame
        ttk.Label(self.scroll_frame, text="Semana", width=16).grid(row=0, column=0, sticky=tk.W)
        ttk.Label(self.scroll_frame, text="Lunes", width=28).grid(row=0, column=1, sticky=tk.W)
        ttk.Label(self.scroll_frame, text="Martes", width=28).grid(row=0, column=2, sticky=tk.W)
        ttk.Label(self.scroll_frame, text="Miércoles (Sesión 1)", width=28).grid(row=0, column=3, sticky=tk.W)
        ttk.Label(self.scroll_frame, text="Miércoles (Sesión 2)", width=28).grid(row=0, column=4, sticky=tk.W)

        # Mapa semana -> (Text lunes, Text martes, Text mié1, Text mié2)
        self.inputs = {}

        exams = self._get_exam_dates()
        for i, wd in enumerate(self.week_dates, start=1):
            row_idx = i

            holiday_mon = wd.lunes in self.holidays
            holiday_tue = wd.martes in self.holidays
            holiday_wed = wd.miercoles in self.holidays

            semana_lbl = f"{wd.semana} ({wd.lunes.strftime('%d/%m')} - {wd.miercoles.strftime('%d/%m')})"
            ttk.Label(self.scroll_frame, text=semana_lbl, width=16).grid(row=row_idx, column=0, padx=(0, 6), sticky=tk.W)

            def mk_text(col, day: date, is_holiday: bool) -> Any:
                t = tk.Text(self.scroll_frame, width=28, height=3, wrap="word")
                t.grid(row=row_idx, column=col, padx=3, sticky=tk.W)
                if is_holiday:
                    name = self.holidays.get(day, "Festivo")
                    t.insert("1.0", f"Festivo: {name}\nNo hay clase")
                    t.config(state=tk.DISABLED)
                elif day in exams:
                    t.insert("1.0", "Examen")
                return t

            txt_mon = mk_text(1, wd.lunes, holiday_mon)
            txt_tue = mk_text(2, wd.martes, holiday_tue)
            txt_w1 = mk_text(3, wd.miercoles, holiday_wed)
            txt_w2 = mk_text(4, wd.miercoles, holiday_wed)
            self.inputs[wd.semana] = (txt_mon, txt_tue, txt_w1, txt_w2)

    def _collect_entries(self) -> Dict[int, Tuple[str, str, str, str]]:
        """Extrae los textos escritos por el usuario, omitiendo celdas bloqueadas por festivo."""
        out: Dict[int, Tuple[str, str, str, str]] = {}
        for semana, (m, t, w1, w2) in self.inputs.items():
            def get_text(txt: Any) -> str:
                if str(txt["state"]) == "disabled":
                    return ""
                return txt.get("1.0", "end").strip()

            out[semana] = (get_text(m), get_text(t), get_text(w1), get_text(w2))
        return out

    def export_excel(self) -> None:
        """Dialoga una ruta y genera el Excel usando build_excel()."""
        title = self.var_title.get().strip() or "Calendario de sesiones"
        subtitle = self.var_sub.get().strip()
        path = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            exams = self._get_exam_dates()
            build_excel(path, title, subtitle, self.week_dates, self._collect_entries(), self.holidays, exams)
            self._save_backup()
            messagebox.showinfo("Listo", f"Archivo Excel generado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el Excel.\n{e}")

    def export_pdf(self) -> None:
        """Dialoga una ruta y genera el PDF si reportlab está instalado."""
        title = self.var_title.get().strip() or "Calendario de sesiones"
        subtitle = self.var_sub.get().strip()
        path = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
        )
        if not path:
            return
        try:
            exams = self._get_exam_dates()
            build_pdf(path, title, subtitle, self.week_dates, self._collect_entries(), self.holidays, exams)
            self._save_backup()
            messagebox.showinfo("Listo", f"Archivo PDF generado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF.\n{e}")

    # ---------- Backup persistence ----------
    def _backup_path(self) -> str:
        return os.path.join(os.path.dirname(__file__), "calendario_backup.json")

    def _read_backup_file(self) -> Optional[Dict[str, Any]]:
        try:
            path = self._backup_path()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            return None
        return None

    def _save_backup(self) -> None:
        try:
            data: Dict[str, Any] = {}
            data["title"] = self.var_title.get().strip()
            data["subtitle"] = self.var_sub.get().strip()
            s = self._get_selected_start_date()
            data["start_date"] = s.isoformat()
            data["weeks"] = int(self.var_weeks.get()) if str(self.var_weeks.get()).isdigit() else self.weeks
            exams = sorted(self._get_exam_dates())
            data["exam_dates"] = [d.isoformat() for d in exams]
            entries = self._collect_entries()
            data["entries"] = {str(k): list(v) for k, v in entries.items()}
            with open(self._backup_path(), "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _apply_saved_entries(self, saved: Dict[str, List[str]]) -> None:
        for semana, texts in saved.items():
            try:
                sem = int(semana)
            except Exception:
                continue
            if sem not in self.inputs:
                continue
            widgets = self.inputs[sem]
            for idx, txt in enumerate(texts[:4]):
                if not txt:
                    continue
                t = widgets[idx]
                if str(t["state"]) == "disabled":
                    continue
                cur = t.get("1.0", "end").strip()
                if cur.startswith("Examen"):
                    t.insert("end", ("\n" if txt else "") + txt)
                else:
                    t.delete("1.0", "end")
                    t.insert("1.0", txt)

    def _load_backup(self) -> None:
        data = self._read_backup_file()
        if not data:
            return
        if isinstance(data.get("title"), str):
            self.var_title.set(data["title"])
        if isinstance(data.get("subtitle"), str):
            self.var_sub.set(data["subtitle"])
        try:
            if isinstance(data.get("start_date"), str):
                d = date.fromisoformat(data["start_date"])
                if TKCAL_OK:
                    self.date_widget.set_date(d)
                else:
                    self._var_day.set(f"{d.day:02d}")
                    self._var_month.set(f"{d.month:02d}")
                    self._var_year.set(str(d.year))
        except Exception:
            pass
        if isinstance(data.get("weeks"), int):
            self.var_weeks.set(str(data["weeks"]))
        ex: List[date] = []
        if isinstance(data.get("exam_dates"), list):
            for s in data["exam_dates"]:
                try:
                    ex.append(date.fromisoformat(s))
                except Exception:
                    pass
        for i, d in enumerate(ex[:len(self.exam_inputs)]):
            w = self.exam_inputs[i]
            try:
                if TKCAL_OK and hasattr(w, 'set_date'):
                    w.set_date(d)
                else:
                    w.delete(0, "end")
                    w.insert(0, d.strftime("%d/%m/%Y"))
            except Exception:
                continue
        # Rebuild now that inputs may have changed
        self.rebuild_calendar()
        entries = data.get("entries")
        if isinstance(entries, dict):
            self._apply_saved_entries(entries)

    def _on_close(self) -> None:
        self._save_backup()
        try:
            self.root.destroy()
        except Exception:
            pass


def run_gui() -> int:
    """Punto de entrada de la GUI; devuelve código de salida (0=OK)."""
    if tk is None:
        print("tkinter no está disponible en este entorno.")
        return 2
    root = tk.Tk()
    CalendarGUI(root)
    root.mainloop()
    return 0


def main(argv: Optional[List[str]] = None) -> int:
    """Entrada principal del módulo.

    Por ahora siempre abre la GUI. En el futuro se podría añadir CLI para
    exportar directamente sin abrir ventana (p. ej. leyendo un JSON/CSV).
    """
    return run_gui()


if __name__ == "__main__":
    sys.exit(main())
