import calendar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Crear un libro de trabajo de Excel y seleccionar la hoja activa
workbook = Workbook()
sheet = workbook.active

# Crear una lista con los nombres de los meses que queremos generar
meses = ["Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre"]

# Escribir los nombres de los meses en la primera fila de la hoja
for i in range(len(meses)):
    sheet.cell(row=1, column=i+1).value = meses[i]

# Generar el calendario mes a mes
for mes in range(3, 10):
    # Obtener el calendario para el mes actual
    cal = calendar.monthcalendar(2023, mes)

    # Escribir el calendario en la hoja
    row = 2
    for week in cal:
        for day in week:
            if day != 0:
                sheet.cell(row=row, column=mes-2).value = day
                row += 1

# Ajustar el ancho de las columnas para que se ajusten al contenido
for col in range(len(meses)):
    col_letter = get_column_letter(col+1)
    sheet.column_dimensions[col_letter].auto_size = True

# Guardar el libro de trabajo en un archivo
workbook.save("calendario_2023.xlsx")